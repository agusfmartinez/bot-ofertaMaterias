"""
UNAHUR SIU Guaraní - Scraper de horarios de cursada
"""

import requests
from bs4 import BeautifulSoup
import json
import re
import time
import os
import sys
from datetime import datetime
from dotenv import load_dotenv

# La consola de Windows usa cp1252 y rompe con los símbolos → ✓ ─ de los prints
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except AttributeError:
    pass

load_dotenv()

# ──────────────────────────────────────────────
#  CONFIGURACIÓN
# ──────────────────────────────────────────────
USUARIO    = os.getenv("USUARIO")
PASSWORD   = os.getenv("PASSWORD")
CARRERA_ID = os.getenv("CARRERA_ID")
GENERA_EXCEL = os.getenv("GENERA_EXCEL", "true").strip().lower() == "true"

BASE_URL    = "https://servicios.unahur.edu.ar/unahur3w"
FILE_ESTADO = "materias_estado.json"
FILE_DATOS  = "horarios_data.json"
FILE_EXCEL  = "horarios_unahur.xlsx"
DELAY       = 1.5

# ──────────────────────────────────────────────
#  PERSISTENCIA
# ──────────────────────────────────────────────

def cargar_json(path, default):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return default

def guardar_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ──────────────────────────────────────────────
#  EXTRAER PAGELETS via BeautifulSoup
# ──────────────────────────────────────────────

def extraer_on_arrival(html_bytes):
    soup = BeautifulSoup(html_bytes, "html.parser", from_encoding="iso-8859-1")
    pagelets = {}
    for script in soup.find_all("script"):
        text = script.get_text()
        if "kernel.renderer.on_arrival" not in text:
            continue
        m = re.search(r'kernel\.renderer\.on_arrival\((\{.*\})\)', text, re.DOTALL)
        if not m:
            continue
        raw = m.group(1)
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            try:
                decoder = json.JSONDecoder()
                data, _ = decoder.raw_decode(raw)
            except Exception:
                continue
        pid     = data.get("info", {}).get("id") or ""
        content = data.get("content", "")
        if pid and content:
            pagelets[pid] = content
    return pagelets

# ──────────────────────────────────────────────
#  SESIÓN HTTP
# ──────────────────────────────────────────────

def crear_sesion():
    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept-Language": "es-AR,es;q=0.9",
    })
    return s

def login(sesion):
    print("[LOGIN] Iniciando sesión...")
    sesion.get(f"{BASE_URL}/acceso", timeout=15)
    r = sesion.post(
        f"{BASE_URL}/acceso?auth=form",
        data={"usuario": USUARIO, "password": PASSWORD, "login": "Ingresar"},
        timeout=15,
        allow_redirects=True,
    )
    r.raise_for_status()
    if "logout" not in r.text and "inicio_alumno" not in r.url:
        raise RuntimeError("Login fallido. Verificar usuario y contraseña.")
    print(f"[LOGIN] OK → {r.url}")

# ──────────────────────────────────────────────
#  OBTENER LISTA DE MATERIAS
# ──────────────────────────────────────────────

def obtener_nombre_carrera(sesion, carrera_id):
    """Nombre de la carrera activa según el SIU. Se usa para nombrar la hoja del Sheet."""
    r = sesion.get(f"{BASE_URL}/inicio_alumno", timeout=15)
    soup = BeautifulSoup(r.content, "html.parser", from_encoding="iso-8859-1")
    for a in soup.find_all("a", attrs={"data-carrera-id": True}):
        if str(a["data-carrera-id"]).strip() == str(carrera_id).strip():
            return a.get("title", "").strip()
    return ""

def obtener_lista_materias(sesion):
    print("[MATERIAS] Obteniendo listado desde /cursada ...")
    r = sesion.get(f"{BASE_URL}/cursada", timeout=15)
    r.raise_for_status()
    time.sleep(DELAY)

    pagelets   = extraer_on_arrival(r.content)
    html_lista = pagelets.get("lista_materias", "")
    soup = BeautifulSoup(html_lista if html_lista else r.content,
                         "html.parser",
                         **({"from_encoding": "iso-8859-1"} if not html_lista else {}))

    links = soup.find_all("a", href=re.compile("elegir_materia"))
    if not links:
        raise RuntimeError("No se encontraron materias. Revisar login.")

    materias = _links_a_materias(links)
    materias.reverse()
    print(f"[MATERIAS] Total encontradas: {len(materias)}")
    return materias

def _links_a_materias(links):
    materias, seen = [], set()
    for a in links:
        href = a.get("href", "")
        if not href or href in seen:
            continue
        seen.add(href)
        title  = a.get("title", "")
        partes = title.rsplit(" - (", 1)
        nombre = partes[0].strip() if partes else title.strip()
        codigo = partes[1].rstrip(")") if len(partes) == 2 else ""
        materias.append({"nombre": nombre, "codigo": codigo, "url": href})
    return materias

# ──────────────────────────────────────────────
#  EXTRAER COMISIONES DE UNA MATERIA
# ──────────────────────────────────────────────

def limpiar(tag):
    return tag.get_text(strip=True) if tag else ""

def extraer_comisiones(sesion, materia):
    print(f"  → {materia['nombre']} ({materia['codigo']})")
    r = sesion.get(materia["url"], timeout=20)
    r.raise_for_status()
    time.sleep(DELAY)

    pagelets  = extraer_on_arrival(r.content)
    html_info = pagelets.get("info_materia", "")

    for html_src in [html_info, None]:
        if html_src is not None:
            soup = BeautifulSoup(html_src, "html.parser")
        else:
            soup = BeautifulSoup(r.content, "html.parser", from_encoding="iso-8859-1")
        ul = soup.find("ul", id="comisiones")
        if ul:
            break
    else:
        print(f"    [WARN] Sin #comisiones")
        return []

    comisiones = []
    for li in ul.find_all("li", recursive=False):
        clases = li.get("class", [])
        if "comision" not in clases:
            continue

        h4     = li.find("h4")
        titulo = limpiar(h4).replace("→", "").replace("\n", " ").strip() if h4 else ""

        turno_span = li.find("span", class_=lambda c: c and "turno" in c.split())
        turno      = limpiar(turno_span).replace("Turno:", "").strip() if turno_span else ""

        # Múltiples horarios y tipos de clase → unir con " - "
        horarios = [
            limpiar(s).replace("Horario:", "").strip()
            for s in li.find_all("span", class_="horario")
        ]
        tipos = [
            limpiar(s).replace("Tipo clase:", "").strip()
            for s in li.find_all("span", class_="tipo_clase")
        ]
        periodicidades = [
            limpiar(s).replace("Periodicidad:", "").strip()
            for s in li.find_all("span", class_="periodicidad")
        ]

        comisiones.append({
            "comision":      titulo,
            "turno":         turno,
            "horario":       " - ".join(horarios),
            "tipo_clase":    " - ".join(tipos),
            "periodicidad":  " - ".join(periodicidades),
        })

    return comisiones

# ──────────────────────────────────────────────
#  EXPORTAR A EXCEL
# ──────────────────────────────────────────────

def exportar_excel(datos, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Horarios UNAHUR 2026"

    headers = ["Materia", "Código", "Comisión", "Turno", "Horario", "Periodicidad", "Tipo Clase"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", start_color="1F4E79")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    widths = [38, 12, 42, 12, 30, 15, 22]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    fill_alt = PatternFill("solid", start_color="DEEAF1")
    row = 2
    for materia in datos:
        comisiones = materia.get("comisiones", [])
        if not comisiones:
            ws.cell(row=row, column=1, value=materia["nombre"])
            ws.cell(row=row, column=2, value=materia["codigo"])
            ws.cell(row=row, column=3, value="(sin comisiones)")
            row += 1
            continue
        for com in comisiones:
            fill = fill_alt if row % 2 == 0 else None
            vals = [
                materia["nombre"],
                materia["codigo"],
                com["comision"],
                com["turno"],
                com["horario"],
                com["periodicidad"],
                com["tipo_clase"],
            ]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.alignment = Alignment(wrap_text=True)
                if fill:
                    c.fill = fill
            row += 1

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"[EXCEL] Guardado: {path} ({row-2} filas)")

# ──────────────────────────────────────────────
#  MAIN
# ──────────────────────────────────────────────

def main():
    print("=" * 55)
    print("  UNAHUR Scraper - Horarios de Cursada 2026")
    print("=" * 55)

    if USUARIO == "TU_USUARIO_AQUI":
        print("\n[ERROR] Completar USUARIO y PASSWORD antes de ejecutar.")
        sys.exit(1)

    estado     = cargar_json(FILE_ESTADO, {"completadas": [], "total": 0})
    datos      = cargar_json(FILE_DATOS,  [])
    procesadas = set(estado.get("completadas", []))

    sesion = crear_sesion()
    login(sesion)

    carrera = obtener_nombre_carrera(sesion, CARRERA_ID)
    print(f"[CARRERA] id={CARRERA_ID} → {carrera or '(nombre no encontrado)'}")

    sesion.get(f"{BASE_URL}/acceso/cambiar_carrera?id={CARRERA_ID}&op=cursada", timeout=15)
    time.sleep(1)

    materias = obtener_lista_materias(sesion)
    estado["total"]   = len(materias)
    estado["carrera"] = carrera
    guardar_json(FILE_ESTADO, estado)

    print(f"\n[INICIO] Total: {len(materias)} | Ya procesadas: {len(procesadas)} | Pendientes: {len(materias)-len(procesadas)}\n")

    for i, materia in enumerate(materias, 1):
        if materia["url"] in procesadas:
            print(f"  [{i}/{len(materias)}] SKIP: {materia['nombre']}")
            continue

        print(f"  [{i}/{len(materias)}]", end=" ")
        try:
            comisiones = extraer_comisiones(sesion, materia)
            datos.append({
                "nombre":       materia["nombre"],
                "codigo":       materia["codigo"],
                "url":          materia["url"],
                "comisiones":   comisiones,
                "scrapeado_en": datetime.now().isoformat(),
            })
            guardar_json(FILE_DATOS, datos)
            procesadas.add(materia["url"])
            estado["completadas"]          = list(procesadas)
            estado["ultima_actualizacion"] = datetime.now().isoformat()
            guardar_json(FILE_ESTADO, estado)
            print(f"    ✓ {len(comisiones)} comisión(es)")
        except Exception as e:
            print(f"    [ERROR] {e}")
            time.sleep(2)

    print(f"\n[DONE] {len(procesadas)}/{len(materias)} materias procesadas")
    print(f"\n  JSON datos:  {FILE_DATOS}")
    print(f"  JSON estado: {FILE_ESTADO}")

    if GENERA_EXCEL:
        exportar_excel(datos, FILE_EXCEL)
        print(f"  Excel:       {FILE_EXCEL}")
    else:
        print("  Excel:       omitido (GENERA_EXCEL=false)")

if __name__ == "__main__":
    main()